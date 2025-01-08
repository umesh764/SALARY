import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import random

# Main application window
root = tk.Tk()
root.title("Employee Salary Calculation")
root.geometry("1200x800")
root.configure(bg="#f2f2f2")

# Employee data for basic pay lookup
employee_basic_pay = {
    "prakash ghodeswar": 10000,
    "sonu sharma": 9500,
    "shriram ninave": 10000,
    "anand kathwate": 20000,
    "kamlesh goyal": 20000,
    "umesh borkar": 12500,
    "manav sur": 20000,
    "ankit khobragade": 10000,
    "anand motghare": 20000,
    "niraj kidey": 10000,
    "prakash nimkar": 12500,
    "anand kumbhare": 12500,
    "chandrakant panchariya": 20000,
    "vijay ulabhaje": 10000,
    "rudesh gedam": 20000,
    "ankush dhawankar": 10000,
    "ramchanda belekar": 10000,
    "ravi pimpalkar": 10000,
    "mahesh salpekar": 10000,
    "kirti chandak": 10000,
    "amit mundada": 20000,
    "govind bhutada": 10000,
    "akash haygune": 9500,
    "sunil katekar": 10000,
    "shrikant paturkar": 12500
}

# List of employees who should not receive PF or have reduced city allowance or gratuity
employees_no_pf = ["shriram ninave", "sunil katekar", "mahesh salpekar", "kirti chandak", "ramchandra belekar", "ravi pimpalkar", "ankush dhawankar"]
employees_reduced_city_allowance = ["prakash ghodeswar", "sonu sharma", "amit mundada"]
employees_no_gratuity = ["sunil katekar", "ramchanda belekar", "mahesh salpekar", "ankush dhawankar", "shriram ninave"]
employees_no_conveyance = ["kirti chandak", "sonu sharma", "sunil katekar", "ramchanda belekar", "mahesh salpekar", "ankush dhawankar"]

# Function to validate numeric inputs
def is_numeric(value):
    try:
        float(value)
        return True
    except ValueError:
        return False

# Function to calculate salary and deductions
def calculate_salary():
    try:
        # Fetch inputs
        emp_name = entry_employee_name.get().strip().lower()  # Ensure no leading/trailing spaces and convert to lowercase
        company_name = entry_company_name.get()
        month = entry_month.get()
        post = entry_post.get()

        # Check if the employee exists in the dictionary (case insensitive)
        basic_salary = employee_basic_pay.get(emp_name, 0)

        if basic_salary == 0:
            messagebox.showerror("Error", f"Employee '{emp_name}' not found! Please check the name entered.")
            return

        # Display Basic Salary
        entry_basic_salary.delete(0, tk.END)
        entry_basic_salary.insert(0, f"{basic_salary:.2f}")

        # Check if days worked is numeric
        if not is_numeric(entry_days_worked.get()):
            messagebox.showerror("Error", "Please enter a valid number of days worked.")
            return

        # Get the number of days worked from the Entry widget and convert to int
        days_worked = int(entry_days_worked.get())  # Total days worked in the month
        total_days_in_month = 26  # Assuming 26 days in a month for simplicity

        # Adjust Basic Salary based on days worked (dynamic basic calculation)
        adjusted_basic_salary = (basic_salary / total_days_in_month) * days_worked

        # Fixed components based on adjusted basic salary
        hra = round(adjusted_basic_salary * 0.40, 2)  # HRA 40% of Adjusted Salary
        conveyance = round(adjusted_basic_salary * 0.20, 2)  # Conveyance 20% of Adjusted Salary

        # Deductions and bonuses
        epf = round(adjusted_basic_salary * 0.13, 2) if emp_name not in employees_no_pf else 0  # EPF 13.25% of Adjusted Salary (skip for some employees)
        provident_fund = round(adjusted_basic_salary * 0.12, 2) if emp_name not in employees_no_pf else 0  # Provident Fund 12% of Adjusted Salary (skip for some employees)
        gratuity = round((adjusted_basic_salary) * 0.0481, 2)  # Gratuity 4.81% of Gross Salary (Basic + HRA + Conveyance)
        
        # Leave Deduction (1.25% of Gross Salary)
        gross_salary = adjusted_basic_salary + hra + conveyance
        leave_deduction = round(gross_salary * 0.10745, 2)  # 1.25% of Gross Salary
        
        # Cap Leave Deduction
        if leave_deduction >  21900:
            leave_deduction = 0

        bonus = round(adjusted_basic_salary * 0.0833, 2)  # Bonus 8.33% of Adjusted Salary

        # Professional Tax
        professional_tax = 200  # Default max cap for professional tax
        if adjusted_basic_salary <= 7500:
            professional_tax = 175  # If basic pay is 7500 or less, PT is fixed at 175
        else:
            professional_tax = min(professional_tax, 200)  # If more, PT is capped at 200

        # Updated City Allowance Calculation (using adjusted salary)
        city_allowance_percentage = random.uniform(1.541, 1.541) if emp_name not in employees_reduced_city_allowance else random.uniform(1.541, 1.541)
        city_allowance = round(adjusted_basic_salary * city_allowance_percentage, 2)  # Calculate based on adjusted salary, not basic salary

        # ESIC (8.05% of Basic Pay)
        esic = round(adjusted_basic_salary * 0.0, 2)

        # Total Emoluments (TEC)
        total_emoluments = gross_salary + city_allowance + bonus + leave_deduction + epf + professional_tax + gratuity 
        total_emoluments_12_months = total_emoluments * 12  # Multiply by 12 for annual figure

        # Total Deductions
        total_deductions = epf + provident_fund + gratuity + leave_deduction + professional_tax + esic
        take_home_salary = gross_salary + city_allowance - total_deductions + provident_fund + bonus - professional_tax + leave_deduction

        # Get the Advance Deduction (new field)
        advance_deduction = 0
        if entry_advance.get().strip():
            if not is_numeric(entry_advance.get()):
                messagebox.showerror("Error", "Please enter a valid numeric value for Advance Deduction.")
                return
            advance_deduction = float(entry_advance.get())

        # Get the TDS Deduction (new field)
        tds_deduction = 0
        if entry_tds.get().strip():
            if not is_numeric(entry_tds.get()):
                messagebox.showerror("Error", "Please enter a valid numeric value for TDS Deduction.")
                return
            tds_deduction = float(entry_tds.get())

        # Subtract the advance deduction and TDS deduction from the take-home salary
        final_take_home_salary = take_home_salary - advance_deduction - tds_deduction

        # Update calculated fields
        entry_basic_rate.delete(0, tk.END)
        entry_basic_rate.insert(0, f"{adjusted_basic_salary:.2f}")

        entry_hra.delete(0, tk.END)
        entry_hra.insert(0, f"{hra:.2f}")

        entry_conveyance.delete(0, tk.END)
        entry_conveyance.insert(0, f"{conveyance:.2f}")

        entry_city_allowance.delete(0, tk.END)
        entry_city_allowance.insert(0, f"{city_allowance:.2f}")

        entry_epf.delete(0, tk.END)
        entry_epf.insert(0, f"{epf:.2f}")

        entry_provident_fund.delete(0, tk.END)
        entry_provident_fund.insert(0, f"{provident_fund:.2f}")

        entry_gratuity.delete(0, tk.END)
        entry_gratuity.insert(0, f"{gratuity:.2f}")

        entry_leave.delete(0, tk.END)
        entry_leave.insert(0, f"{leave_deduction:.2f}")

        entry_bonus.delete(0, tk.END)
        entry_bonus.insert(0, f"{bonus:.2f}")

        entry_professional_tax.delete(0, tk.END)
        entry_professional_tax.insert(0, f"{professional_tax:.2f}")

        entry_take_home.delete(0, tk.END)
        entry_take_home.insert(0, f"{final_take_home_salary:.2f}")

        entry_total_emoluments.delete(0, tk.END)
        entry_total_emoluments.insert(0, f"{total_emoluments_12_months:.2f}")

        # Display ESIC value
        entry_esic.delete(0, tk.END)
        entry_esic.insert(0, f"{esic:.2f}")

        # Prompt to export data
        if messagebox.askyesno("Export", "Do you want to save this data to an Excel file?"):
            save_to_excel(company_name, month, post, emp_name, adjusted_basic_salary, hra, conveyance, city_allowance, epf, provident_fund, gratuity, leave_deduction, bonus, professional_tax, esic, final_take_home_salary, total_emoluments_12_months, advance_deduction, tds_deduction)
    except Exception as e:
        messagebox.showerror("Error", f"Calculation error: {str(e)}")

# Function to save data to Excel
def save_to_excel(company_name, month, post, emp_name, basic_rate, hra, conveyance, city_allowance, epf, provident_fund, gratuity, leave_deduction, bonus, professional_tax, esic, take_home_salary, total_emoluments_12_months, advance_deduction, tds_deduction):
    wb = Workbook()
    ws = wb.active
    ws.title = "Salary Details"

    # Define header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Add headers
    headers = ["Field", "Value"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Add employee data
    data = [
        ("Company Name", company_name),
        ("Month", month),
        ("Post", post),
        ("Employee Name", emp_name),
        ("Basic Rate (Calculated)", basic_rate),
        ("HRA (40% of Basic Pay)", hra),
        ("Conveyance (20% of Basic Pay)", conveyance),
        ("City Allowance", city_allowance),
        ("EPF Contribution", epf),
        ("Provident Fund Contribution", provident_fund),
        ("Gratuity", gratuity),
        ("Leave Deduction", leave_deduction),
        ("Bonus", bonus),
        ("Professional Tax", professional_tax),
        ("ESIC (8.05% of Basic Pay)", esic),
        ("Advance Deduction", advance_deduction),
        ("TDS Deduction", tds_deduction),
        ("Take Home Salary", take_home_salary),
        ("Total Emoluments (TEC)", total_emoluments_12_months)
    ]

    for row_num, (field, value) in enumerate(data, 2):
        ws.cell(row=row_num, column=1, value=field)
        ws.cell(row=row_num, column=2, value=value)

    file_name = f"{emp_name}_Salary_Details.xlsx"
    wb.save(file_name)
    messagebox.showinfo("Success", f"File saved as {file_name}")

# Create labeled input fields for left side
def create_label_entry_left(label_text, row, col):
    label = tk.Label(root, text=label_text, bg="#f2f2f2", font=("Arial", 10, "bold"))
    label.grid(row=row, column=col, padx=10, pady=5, sticky="w")
    entry = tk.Entry(root, font=("Arial", 10), width=30)
    entry.grid(row=row, column=col+1, padx=10, pady=5)
    return entry

# Create labeled input fields for right side
def create_label_entry_right(label_text, row, col):
    label = tk.Label(root, text=label_text, bg="#f2f2f2", font=("Arial", 10, "bold"))
    label.grid(row=row, column=col, padx=10, pady=5, sticky="w")
    entry = tk.Entry(root, font=("Arial", 10), width=30)
    entry.grid(row=row, column=col+1, padx=10, pady=5)
    return entry

# Initialize input fields
entry_employee_name = create_label_entry_left("Employee Name:", 0, 0)
entry_company_name = create_label_entry_left("Company Name:", 1, 0)
entry_month = create_label_entry_left("Month:", 2, 0)
entry_post = create_label_entry_left("Post:", 3, 0)
entry_days_worked = create_label_entry_left("Working Days:", 4, 0)

entry_basic_salary = create_label_entry_right("Basic Salary:", 0, 2)
entry_basic_rate = create_label_entry_right("Basic Rate (Calculated):", 1, 2)
entry_hra = create_label_entry_right("HRA (40% of Basic Pay):", 2, 2)
entry_conveyance = create_label_entry_right("Conveyance (20% of Basic Pay):", 3, 2)
entry_city_allowance = create_label_entry_right("City Allowance:", 4, 2)
entry_epf = create_label_entry_right("EPF Contribution:", 5, 2)
entry_provident_fund = create_label_entry_right("Provident Fund Contribution:", 6, 2)
entry_gratuity = create_label_entry_right("Gratuity:", 7, 2)
entry_leave = create_label_entry_right("Leave Deduction:", 8, 2)
entry_bonus = create_label_entry_right("Bonus:", 9, 2)
entry_professional_tax = create_label_entry_right("Professional Tax:", 10, 2)
entry_take_home = create_label_entry_right("Take Home Salary:", 11, 2)
entry_total_emoluments = create_label_entry_right("Total Emoluments (TEC):", 12, 2)
entry_esic = create_label_entry_right("ESIC:", 13, 2)

# Add advance deduction and TDS fields
entry_advance = create_label_entry_left("Advance Deduction:", 14, 0)
entry_tds = create_label_entry_left("TDS Deduction:", 15, 0)

# Button to calculate salary
button_calculate = tk.Button(root, text="Calculate Salary", font=("Arial", 12, "bold"), command=calculate_salary, bg="#4CAF50", fg="white")
button_calculate.grid(row=16, column=0, columnspan=4, pady=20)

root.mainloop()
